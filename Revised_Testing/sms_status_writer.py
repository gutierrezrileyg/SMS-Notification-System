import json
import os
import re
import threading
from datetime import datetime
from flask import Flask, jsonify, request
import openpyxl

app = Flask(__name__)

# Configurable Constants - matches your Excel sheet file name
EXCEL_FILE = "Testing Interface (1).xlsx"
QUEUE_FILE = "pending_updates.json"
WEBHOOK_SECRET = "CENRO_SECRET_2026"

excel_lock = threading.Lock()
STATUS_PRIORITY = {"pending": 1, "retrying": 2, "sent": 3, "failed": 3}

def normalize_phone(phone):
    digits = re.sub(r"\D", "", str(phone or ""))
    if digits.startswith("0") and len(digits) == 11:
        digits = "63" + digits[1:]
    return digits

def apply_update_to_sheet(ws, msg_id, phone, new_status, timestamp):
    phone_str = normalize_phone(phone)
    new_status_clean = str(new_status or "").strip().lower()
    new_priority = STATUS_PRIORITY.get(new_status_clean, 0)

    found_row = None
    current_status = None

    for row in range(4, ws.max_row + 1):
        cell_phone = normalize_phone(ws.cell(row=row, column=5).value)
        if phone_str and cell_phone == phone_str:
            found_row = row
            current_status = str(ws.cell(row=row, column=7).value or "").strip().lower()
            break

    current_priority = STATUS_PRIORITY.get(current_status, 0)

    if found_row:
        if new_priority >= current_priority:
            ws.cell(row=found_row, column=7, value=new_status_clean.capitalize())
            ws.cell(row=found_row, column=3, value=timestamp)
    else:
        for row in range(4, ws.max_row + 5):
            if not ws.cell(row=row, column=5).value:
                ws.cell(row=row, column=5, value=phone_str)
                ws.cell(row=row, column=7, value=new_status_clean.capitalize())
                ws.cell(row=row, column=3, value=timestamp)
                break

    return True

def save_to_queue(data):
    queue = []
    if os.path.exists(QUEUE_FILE):
        try:
            with open(QUEUE_FILE, "r") as f:
                queue = json.load(f)
        except Exception:
            queue = []

    queue.append(data)
    with open(QUEUE_FILE, "w") as f:
        json.dump(queue, f, indent=2)

def process_pending_queue(ws):
    if not os.path.exists(QUEUE_FILE):
        return

    try:
        with open(QUEUE_FILE, "r") as f:
            queue = json.load(f)

        if not queue:
            return

        remaining_queue = []
        for item in queue:
            success = apply_update_to_sheet(
                ws,
                item["message_id"],
                item["phone"],
                item["status"],
                item["timestamp"],
            )
            if not success:
                remaining_queue.append(item)

        if remaining_queue:
            with open(QUEUE_FILE, "w") as f:
                json.dump(remaining_queue, f)
        else:
            os.remove(QUEUE_FILE)
    except Exception as e:
        print(f"Error processing pending queue: {e}")

def async_update_excel(data):
    with excel_lock:
        try: # Fixed typo from 'lry:' to 'try:'
            if not os.path.exists(EXCEL_FILE):
                print(f"Error: Target Excel file '{EXCEL_FILE}' not found. Please create it first.")
                return

            wb = openpyxl.load_workbook(EXCEL_FILE)
            
            if "Database" in wb.sheetnames:
                ws = wb["Database"]
            else:
                ws = wb.active

            process_pending_queue(ws)

            apply_update_to_sheet(
                ws,
                data["message_id"],
                data["phone"],
                data["status"],
                data["timestamp"],
            )

            wb.save(EXCEL_FILE)
        except PermissionError:
            save_to_queue(data)
        except Exception as e:
            print(f"Background write error: {e}")

@app.route("/unisms-webhook", methods=["POST"])
def handle_webhook():
    if request.headers.get("X-Webhook-Secret") != WEBHOOK_SECRET:
        return jsonify({"error": "Unauthorized"}), 401

    payload = request.json or {}
    
    msg_id = payload.get("id") or payload.get("message_id")
    message_obj = payload.get("message") or {}
    phone = payload.get("phone") or message_obj.get("recipient")
    status = payload.get("status") or message_obj.get("status")

    if not status or not phone:
        return jsonify({"error": "Invalid payload format"}), 400

    update_data = {
        "message_id": msg_id,
        "phone": phone,
        "status": status,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    threading.Thread(target=async_update_excel, args=(update_data,)).start()

    return jsonify({"message": "Webhook received successfully"}), 200

if __name__ == "__main__":
    if not os.path.exists(EXCEL_FILE):
        print(f"Warning: '{EXCEL_FILE}' does not exist yet. Please place your template file in the directory.")
    app.run(port=5000, debug=True)