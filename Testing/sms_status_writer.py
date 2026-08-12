import json
import os
import re
import threading
from datetime import datetime
from flask import Flask, jsonify, request
import openpyxl

app = Flask(__name__)

# Configurable Constants
EXCEL_FILE = "CENRO_SMS_Database.xlsx"
QUEUE_FILE = "pending_updates.json"
WEBHOOK_SECRET = "CENRO_SECRET_2026"  # Must match the header token set in UniSMS

# Thread safety lock for concurrent requests
excel_lock = threading.Lock()

# Status priority rules (prevents out-of-order webhook delivery overwrites)
STATUS_PRIORITY = {"pending": 1, "retrying": 2, "sent": 3, "failed": 3}


def normalize_phone(phone):
    """Converts phone numbers (+639..., 09..., 639...) to a standardized 12-digit string."""
    digits = re.sub(r"\D", "", str(phone or ""))
    if digits.startswith("0") and len(digits) == 11:
        digits = "63" + digits[1:]
    return digits


def apply_update_to_sheet(ws, msg_id, phone, new_status, timestamp):
    """Finds matching row based on Phone Number (Column E) and updates Status (Column G) and Sent at (Column C)."""
    phone_str = normalize_phone(phone)
    new_status_clean = str(new_status or "").strip().lower()
    new_priority = STATUS_PRIORITY.get(new_status_clean, 0)

    found_row = None
    current_status = None

    # Loop through the table rows starting from row 4 down to max_row
    for row in range(4, ws.max_row + 1):
        cell_phone = normalize_phone(ws.cell(row=row, column=5).value) # Column E is Phone Numbers

        if phone_str and cell_phone == phone_str:
            found_row = row
            current_status = str(ws.cell(row=row, column=7).value or "").strip().lower() # Column G is Status
            break

    current_priority = STATUS_PRIORITY.get(current_status, 0)

    if found_row:
        # Only update if the new status has equal or higher priority
        if new_priority >= current_priority:
            ws.cell(row=found_row, column=7, value=new_status_clean.capitalize()) # Update Status (Column G)
            ws.cell(row=found_row, column=3, value=timestamp) # Update Sent at (Column C)
    else:
        # Optional: If phone number doesn't exist in the pre-made table, find the next empty row in Column E
        for row in range(4, ws.max_row + 5):
            if not ws.cell(row=row, column=5).value:
                ws.cell(row=row, column=5, value=phone_str)
                ws.cell(row=row, column=7, value=new_status_clean.capitalize())
                ws.cell(row=row, column=3, value=timestamp)
                break

    return True


def save_to_queue(data):
    """Saves update to JSON fallback file if Excel is locked by a user."""
    queue = []
    if os.path.exists(QUEUE_FILE):
        try:
            with open(QUEUE_FILE, "r") as f:
                queue = json.load(f)
        except Exception:
            queue = []

    queue.append(data)
    with open(QUEUE_FILE, "w") as f:
        json.dump(queue, f, indent=2)


def process_pending_queue(ws):
    """Flushes previously saved JSON updates into Excel once unlocked."""
    if not os.path.exists(QUEUE_FILE):
        return

    try:
        with open(QUEUE_FILE, "r") as f:
            queue = json.load(f)

        if not queue:
            return

        remaining_queue = []
        for item in queue:
            success = apply_update_to_sheet(
                ws,
                item["message_id"],
                item["phone"],
                item["status"],
                item["timestamp"],
            )
            if not success:
                remaining_queue.append(item)

        if remaining_queue:
            with open(QUEUE_FILE, "w") as f:
                json.dump(remaining_queue, f)
        else:
            os.remove(QUEUE_FILE)
    except Exception as e:
        print(f"Error processing pending queue: {e}")


def async_update_excel(data):
    """Executes file operations in a background thread to prevent API timeouts."""
    with excel_lock:
        try:
            if not os.path.exists(EXCEL_FILE):
                print(f"Error: Target Excel file '{EXCEL_FILE}' not found. Please create it first.")
                return

            wb = openpyxl.load_workbook(EXCEL_FILE)
            
            # Target your specific sheet tab name ("Database")
            if "Database" in wb.sheetnames:
                ws = wb["Database"]
            else:
                ws = wb.active

            # Flush any queued updates from previous file locks
            process_pending_queue(ws)

            # Apply current update
            apply_update_to_sheet(
                ws,
                data["message_id"],
                data["phone"],
                data["status"],
                data["timestamp"],
            )

            wb.save(EXCEL_FILE)
        except PermissionError:
            # Excel is open in desktop application -> Save to JSON fallback queue
            save_to_queue(data)
        except Exception as e:
            print(f"Background write error: {e}")


@app.route("/unisms-webhook", methods=["POST"])
def handle_webhook():
    # Security Verification
    if request.headers.get("X-Webhook-Secret") != WEBHOOK_SECRET:
        return jsonify({"error": "Unauthorized"}), 401

    payload = request.json or {}
    
    # Extract fields based on standard UniSMS webhook payload structure
    msg_id = payload.get("id") or payload.get("message_id")
    message_obj = payload.get("message") or {}
    phone = payload.get("phone") or message_obj.get("recipient")
    status = payload.get("status") or message_obj.get("status")

    if not status or not phone:
        return jsonify({"error": "Invalid payload format"}), 400

    update_data = {
        "message_id": msg_id,
        "phone": phone,
        "status": status,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    # Dispatch write job asynchronously
    threading.Thread(target=async_update_excel, args=(update_data,)).start()

    # Immediately notify UniSMS with 200 OK
    return jsonify({"message": "Webhook received successfully"}), 200


if __name__ == "__main__":
    if not os.path.exists(EXCEL_FILE):
        print(f"Warning: '{EXCEL_FILE}' does not exist yet. Please place your template file in the directory.")
    app.run(port=5000, debug=True)