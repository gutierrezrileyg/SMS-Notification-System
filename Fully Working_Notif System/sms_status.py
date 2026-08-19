import io
import site
import sqlite3
import urllib.request
import random
from dataclasses import asdict, dataclass
from typing import Any, Dict, List, Optional
import pandas as pd
import openpyxl  # Required for updating the Excel file in-place

# ==============================================================================
# 1. RESIDENT PROFILE SCHEMA (Dataclass & SQLite Schema)
# ==============================================================================
@dataclass
class ResidentProfile:
    """Schema for individual resident profiles containing personal data,
    contact details, and location metadata.
    """

    resident_id: Optional[int] = None
    full_name: str = "Unknown"
    age: Optional[int] = None
    phone_number: str = ""
    email: Optional[str] = None
    # Location Metadata
    barangay: str = ""
    city: str = "Sto Tomas City"
    province: str = "Batangas"
    address_line: Optional[str] = None


class ResidentDatabase:
    """SQLite Database Manager to persist and query Resident Profiles."""

    def __init__(self, db_name: str = ":memory:"):
        self.conn = sqlite3.connect(db_name)
        self.create_table()

    def create_table(self):
        """Creates the resident profiles table schema in SQLite."""
        schema = """
        CREATE TABLE IF NOT EXISTS resident_profiles (
            resident_id INTEGER PRIMARY KEY AUTOINCREMENT,
            full_name TEXT NOT NULL,
            age INTEGER,
            phone_number TEXT NOT NULL,
            email TEXT,
            barangay TEXT NOT NULL,
            city TEXT NOT NULL,
            province TEXT NOT NULL,
            address_line TEXT,
            UNIQUE(phone_number, barangay)
        );
        """
        with self.conn:
            self.conn.execute(schema)

    def insert_profile(self, profile: ResidentProfile) -> bool:
        """Inserts a ResidentProfile dataclass into the SQLite table."""
        query = """
        INSERT OR IGNORE INTO resident_profiles
        (full_name, age, phone_number, email, barangay, city, province, address_line)
        VALUES (:full_name, :age, :phone_number, :email, :barangay, :city, :province, :address_line);
        """
        with self.conn:
            cursor = self.conn.execute(query, asdict(profile))
            return cursor.rowcount > 0

    def get_profiles_by_barangay(
        self, barangay_name: str
    ) -> List[ResidentProfile]:
        """Fetches all valid resident profiles for a given barangay."""
        query = """
        SELECT resident_id, full_name, age, phone_number, email, barangay, city, province, address_line
        FROM resident_profiles
        WHERE LOWER(barangay) = LOWER(?);
        """
        cursor = self.conn.cursor()
        cursor.execute(query, (barangay_name.strip(),))
        rows = cursor.fetchall()

        return [
            ResidentProfile(
                resident_id=row[0],
                full_name=row[1],
                age=row[2],
                phone_number=row[3],
                email=row[4],
                barangay=row[5],
                city=row[6],
                province=row[7],
                address_line=row[8],
            )
            for row in rows
        ]


# ==============================================================================
# 2. RESIDENT DATA FETCHER (Excel & CSV Loader)
# ==============================================================================
class ResidentDataFetcher:
    """Retrieves resident dataset from local or remote Excel (.xlsx) / CSV files."""

    def __init__(self, excel_url_or_path: str):
        self.source = excel_url_or_path
        self.df = self._load_dataframe(excel_url_or_path)
        self._standardize_columns()

    def _load_dataframe(self, path_or_url: str) -> pd.DataFrame:
        """Loads dataset from online URLs or local files using Pandas."""
        try:
            if path_or_url.endswith(".xlsx") or path_or_url.endswith(".xls"):
                df = pd.read_excel(
                    path_or_url,
                    sheet_name="Database",
                    header=2,
                    engine="openpyxl",
                )
            else:
                df = pd.read_csv(path_or_url, header=2)

            if "Barangay" in df.columns:
                df["Barangay"] = df["Barangay"].ffill()

            return df

        except FileNotFoundError:
            print(f"[File Error] Could not find spreadsheet at: {path_or_url}")
            return pd.DataFrame()
        except Exception as e:
            print(f"[System Error] Failed to read spreadsheet: {e}")
            return pd.DataFrame()

    def _standardize_columns(self):
        """Standardizes flexible spreadsheet column headers."""
        if self.df.empty:
            return

        col_mapping = {}
        for col in self.df.columns:
            clean_col = str(col).strip().lower().replace("_", " ")
            if "name" in clean_col:
                col_mapping[col] = "Full_Name"
            elif "age" in clean_col:
                col_mapping[col] = "Age"
            elif any(
                term in clean_col
                for term in ["phone", "mobile", "contact", "cellphone"]
            ):
                col_mapping[col] = "Phone_Number"
            elif "email" in clean_col:
                col_mapping[col] = "Email"
            elif "barangay" in clean_col:
                col_mapping[col] = "Barangay"
            elif "address" in clean_col or "street" in clean_col:
                col_mapping[col] = "Address_Line"

        self.df.rename(columns=col_mapping, inplace=True)

    def extract_raw_profiles(self) -> List[Dict[str, Any]]:
        """Converts raw dataframe rows into dictionary profile records."""
        if self.df.empty:
            return []
        return self.df.to_dict(orient="records")


# ==============================================================================
# 3. DATA HYGIENE (Validation)
# ==============================================================================
def validate_phone(phone: Any) -> bool:
    """Validates if a phone number is present and non-empty."""
    if pd.isna(phone) or not phone:
        return False
    return len(str(phone).strip()) > 0


def sanitize_profile(
    raw_record: Dict[str, Any],
) -> Optional[ResidentProfile]:
    """Cleans, validates, and constructs a ResidentProfile dataclass."""
    raw_phone = raw_record.get("Phone_Number")

    if not validate_phone(raw_phone):
        return None

    phone_number = str(raw_phone).strip()

    raw_age = raw_record.get("Age")
    try:
        age = int(raw_age) if pd.notna(raw_age) else None
    except (ValueError, TypeError):
        age = None

    raw_barangay = str(raw_record.get("Barangay", "")).strip()

    return ResidentProfile(
        full_name=str(
            raw_record.get("Full_Name", "Unknown Resident")
        ).strip(),
        age=age,
        phone_number=phone_number,
        email=(
            str(raw_record.get("Email")).strip()
            if pd.notna(raw_record.get("Email"))
            else None
        ),
        barangay=raw_barangay,
        address_line=(
            str(raw_record.get("Address_Line")).strip()
            if pd.notna(raw_record.get("Address_Line"))
            else None
        ),
    )


# ==============================================================================
# 4. MESSAGE QUEUE & GATEWAY ROUTING LOGIC
# ==============================================================================
def routing_logic(
    barangay_name: str, profiles: List[ResidentProfile]
) -> Dict[str, Any]:
    """Prepares payload from resident profile variables for the message queue."""
    phone_list = [p.phone_number for p in profiles]
    return {
        "destination": "GatewayAPI",
        "barangay": barangay_name,
        "queue_name": f"SMS_QUEUE_{barangay_name.upper().replace(' ', '_')}",
        "total_messages": len(phone_list),
        "payload": phone_list,
        "profiles_count": len(profiles),
    }


def gateway_api(route_data: Dict[str, Any]) -> Dict[str, Any]:
    """Simulates pushing phone payload to the gateway API queue."""
    print(
        f"[Message Queue] Enqueued {route_data['total_messages']} numbers for '{route_data['barangay']}' -> {route_data['queue_name']}"
    )
    return {
        "status": "ready",
        "barangay": route_data["barangay"],
        "records": route_data["payload"],
    }


# ==============================================================================
# 5. EXCEL STATUS UPDATER
# ==============================================================================
class ExcelStatusUpdater:
    """Updates the original Excel file with new statuses from the website."""
    
    def __init__(self, excel_path: str, sheet_name: str = "Database", header_row: int = 3):
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.header_row = header_row
        
    def update_statuses(self, status_updates: Dict[str, str]):
        """Writes the status dictionary back to the Excel file."""
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb[self.sheet_name] if self.sheet_name in wb.sheetnames else wb.active

            phone_col_idx = None
            status_col_idx = None

            for col in range(1, ws.max_column + 2):
                cell_value = str(ws.cell(row=self.header_row, column=col).value).strip().lower()
                
                if any(term in cell_value for term in ["phone", "mobile", "contact", "cellphone"]):
                    phone_col_idx = col
                if cell_value == "status":
                    status_col_idx = col

            if not phone_col_idx:
                print("[Excel Error] Could not find a Phone/Mobile column.")
                return

            if not status_col_idx:
                status_col_idx = ws.max_column + 1
                ws.cell(row=self.header_row, column=status_col_idx, value="Status")

            updates_made = 0
            for row in range(self.header_row + 1, ws.max_row + 1):
                phone_cell_val = str(ws.cell(row=row, column=phone_col_idx).value).strip()
                
                if phone_cell_val in status_updates:
                    ws.cell(row=row, column=status_col_idx, value=status_updates[phone_cell_val])
                    updates_made += 1

            wb.save(self.excel_path)
            print(f"[Success] Updated {updates_made} records in Excel file -> '{self.excel_path}'")

        except PermissionError:
            print(f"[File Error] Cannot write to Excel. Ensure '{self.excel_path}' is CLOSED.")
        except Exception as e:
            print(f"[System Error] Failed to update Excel statuses: {e}")


# ==============================================================================
# MAIN PROCESSING PIPELINE
# ==============================================================================
if __name__ == "__main__":
    db = ResidentDatabase()

    # Updated with your absolute file path
    FILE_PATH = r"C:\Users\Riley G. Gutierrez\OneDrive\Documents\My Codes\unisms-sender\Testing Interface (1).xlsx"
    fetcher = ResidentDataFetcher(FILE_PATH)

    target_barangays = [
        "Poblacion 1",
        "Poblacion 2",
        "Poblacion 3",
        "Poblacion 4",
    ]

    raw_records = fetcher.extract_raw_profiles()

    for record in raw_records:
        profile = sanitize_profile(record)
        if profile and profile.barangay:
            db.insert_profile(profile)

    barangay_profile_variables: Dict[str, List[ResidentProfile]] = {}

    for barangay in target_barangays:
        profiles = db.get_profiles_by_barangay(barangay)
        barangay_profile_variables[barangay] = profiles

    for barangay, profiles_list in barangay_profile_variables.items():
        route = routing_logic(barangay, profiles_list)
        gateway_api(route)

    all_processed_phones = []

    for barangay, profiles_list in barangay_profile_variables.items():
        if profiles_list:
            for profile in profiles_list:
                all_processed_phones.append(profile.phone_number)

    # Simulated status response mapping for the phone numbers
    latest_statuses = {phone: "Sent" for phone in all_processed_phones}

    if latest_statuses:
        excel_updater = ExcelStatusUpdater(excel_path=FILE_PATH)
        excel_updater.update_statuses(latest_statuses)

    print("\nPipeline execution complete.")